from selenium import webdriver
from bs4 import BeautifulSoup
import time
import csv
import tkinter as tk
from tkinter import ttk, messagebox
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import threading
from tkinter import filedialog
import os
import platform
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO
import requests


BASE_URL = 'https://www.carsales.com.au'
BASE_URL_TEMPLATE = f'{BASE_URL}/cars/{{make}}/{{model}}/{{badge}}-badge/{{state}}-state/?page='
YEAR_RANGE_URL_TEMPLATE = f'{BASE_URL}/cars/?q=(And.(C.Make.{{make}}._.(C.Model.{{model}}._.Badge.{{badge}}.))_.State.{{state}}._.Year.range({{min_year}}..{{max_year}}).)'
NUM_PAGES = 1  # Default value for number of pages to scrape

def manual_selection():
    global driver

    # Using the Service object to start the chromedriver
    s = Service(ChromeDriverManager().install())

    # Creating a ChromeOptions object and adding the argument to ignore certificate errors
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')

    # Pass the options to the Chrome webdriver
    driver = webdriver.Chrome(service=s, options=options)

    driver.get(BASE_URL)
    messagebox.showinfo("Info",
                        "Please select make, model, badge, etc., manually on the browser and then come back to the GUI and hit 'Scrape Now'. Do not close the browser.")
def start_scraping_thread():
    scrape_thread = threading.Thread(target=start_scraping)
    scrape_thread.start()

def open_file(filepath):
    """Open a file with its default application."""
    if platform.system() == "Windows":
        os.startfile(filepath)
    elif platform.system() == "Darwin":  # macOS
        os.system(f"open {filepath}")
    else:  # Linux and other UNIX-like systems
        os.system(f"xdg-open {filepath}")

def save_to_csv(data, filepath):
    keys = data[0].keys()
    with open(filepath, 'w', newline='', encoding='utf-8') as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=keys)
        writer.writeheader()
        writer.writerows(data)


def save_to_excel(data, filepath):
    if not data:
        print("No data to save!")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Flatten data and expand dictionaries for better Excel representation
    flat_data = []
    for item in data:
        flat_item = item.copy()
        for key, value in item['KeyDetails'].items():
            flat_item[key] = value
        flat_data.append(flat_item)

    headers = flat_data[0].keys()

    # Write headers to the Excel sheet
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    for row_num, item in enumerate(flat_data, 2):
        for col_num, key in enumerate(headers, 1):
            cell_value = item[key]
            if key == 'Images':
                img_url = item['Images'][0] if item['Images'] else None
                if img_url:
                    response = requests.get(img_url)
                    image = Image(BytesIO(response.content))
                    sheet.column_dimensions[get_column_letter(col_num)].width = 15  # Set column width
                    image.width = 120  # Set image width
                    image.height = 80  # Set image height
                    sheet.row_dimensions[row_num].height = image.height  # Set row height based on image

                    sheet.add_image(image, get_column_letter(col_num) + str(row_num))

            else:
                sheet.cell(row=row_num, column=col_num).value = str(cell_value)

    workbook.save(filepath)

def start_scraping():
    global NUM_PAGES
    if 'driver' not in globals() or not driver.service.process:
        messagebox.showerror("Error", "Please select the car details manually in the browser first.")
        return

    # Checking the value of the spinbox
    pages_value = pages_spinbox.get()
    if not pages_value.isdigit():
        messagebox.showerror("Error", "Invalid number of pages. Please enter a valid number.")
        return

    NUM_PAGES = int(pages_value)

    for page_num in range(1, NUM_PAGES + 1):
        if page_num == 1:
            page_source = driver.page_source
        else:
            driver.get(driver.current_url.split('?')[0] + '?page=' + str(page_num))
            time.sleep(5)
            try:
                page_source = driver.page_source
            except Exception as e:
                print(f"Error occurred: {e}")
                return

        progress_value = (page_num / NUM_PAGES) * 100
        progress_var.set(progress_value)
        app.update_idletasks()
        soup = BeautifulSoup(page_source, 'html.parser')

    car_listings = soup.find_all('div', class_='listing-item')
    car_data = []

    for listing in car_listings:
        car_info = {}
        car_info['NetworkID'] = listing.get('data-webm-networkid')
        car_info['Category'] = listing.get('data-webm-vehcategory')
        car_info['BodyStyle'] = listing.get('data-webm-bodystyle')
        car_info['Make'] = listing.get('data-webm-make')
        car_info['Model'] = listing.get('data-webm-model')
        car_info['State'] = listing.get('data-webm-state')
        car_info['Price'] = listing.get('data-webm-price')
        badge_tag = listing.find('span', class_='csn-badge-title')
        car_info['Badge'] = badge_tag.text.strip() if badge_tag else None
        car_name_tag = listing.find("a", class_="js-encode-search")
        if car_name_tag:
            car_info['Name'] = car_name_tag.text.strip()
            car_info['Link'] = car_name_tag['href']
        images = []
        for img in listing.select(".carousel-item img"):
            img_src = img.get('src') or img.get('data-src')
            if img_src:
                images.append(img_src)
        car_info['Images'] = images
        key_details = {}
        for li in listing.select(".key-details li"):
            label = li.get('data-type')
            if label:
                key_details[label] = li.text.strip()
        car_info['KeyDetails'] = key_details
        car_data.append(car_info)

    if price_filter_var.get() == 1:
        car_data = [car for car in car_data if car.get('Badge') in ['Well below market price', 'Below market price']]

    if car_data:
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        # If user cancels the file dialog, don't proceed
        if not filepath:
            return

        # Use save_to_excel function instead of save_to_csv
        save_to_excel(car_data, filepath)

        # Ask the user if they want to open the saved file
        open_file_question = messagebox.askyesno("Success",
                                                 f"Data has been saved to {filepath}. Do you want to open it?")
        if open_file_question:
            open_file(filepath)
    else:
        messagebox.showerror("Error", "No data found. Please check your selections and try again.")

def center_window(win):
    """Center the tkinter window on the screen"""
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry('{}x{}+{}+{}'.format(width, height, x, y))

app = tk.Tk()
app.title("Car Data Scraper")

# Set an icon for the application (if you have one)
# app.iconbitmap('path_to_icon.ico')
progress_label = ttk.Label(app, text="Scraping Progress:")
progress_label.grid(row=5, column=0, padx=20, pady=10, sticky="w")

progress_var = tk.DoubleVar()  # This variable will store the current progress value.
progress_bar = ttk.Progressbar(app, orient='horizontal', length=300, mode='determinate', variable=progress_var)
progress_bar.grid(row=6, column=0, padx=20, pady=10)

# Style
style = ttk.Style(app)
style.configure('TButton', font=('Arial', 12), padding=10)
style.configure('TCheckbutton', font=('Arial', 12), padding=5)

price_filter_var = tk.IntVar()
price_filter_checkbox = ttk.Checkbutton(app, text="Well under market price", variable=price_filter_var)
price_filter_checkbox.grid(row=0, column=0, padx=20, pady=10, sticky="w")

pages_label = ttk.Label(app, text="Number of Pages to Scrape:")
pages_label.grid(row=1, column=0, padx=20, pady=10, sticky="w")

pages_spinbox = ttk.Spinbox(app, from_=1, to=100)
pages_spinbox.grid(row=2, column=0, padx=20, pady=10)

manual_selection_button = ttk.Button(app, text="Manual Selection", command=manual_selection)
manual_selection_button.grid(row=3, column=0, padx=20, pady=10)

scrape_button = ttk.Button(app, text="Scrape Now", command=start_scraping_thread)
scrape_button.grid(row=4, column=0, padx=20, pady=10)

# Menu bar
menu = tk.Menu(app)
app.config(menu=menu)

file_menu = tk.Menu(menu)
menu.add_cascade(label="File", menu=file_menu)
file_menu.add_command(label="Exit", command=app.quit)

help_menu = tk.Menu(menu)
menu.add_cascade(label="Help", menu=help_menu)
help_menu.add_command(label="About", command=lambda: messagebox.showinfo("About", "Car Data Scraper v1.0"))

# Position the window to the center of the screen
center_window(app)

app.mainloop()